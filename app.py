import streamlit as st
import openpyxl
import csv
import io
import re
import zipfile
from collections import OrderedDict

# ─── Page config ──────────────────────────────────────────────────────────────
st.set_page_config(
    page_title="Eyewa Listing Automation",
    page_icon="👓",
    layout="wide",
)

st.title("👓 Eyewa Listing Automation")
st.caption("Fill Trendyol & Noon templates from SKU codes in seconds.")

# ─── Attribute mappings ───────────────────────────────────────────────────────
DEPT_MAP = {
    "women": "Women", "female": "Women",
    "men": "Men", "male": "Men",
    "unisex": "Unisex",
    "kids": "Kids Unisex", "boys": "Boys", "girls": "Girls",
}
FRAME_COLOR_FAMILY = {
    "black": "Black", "white": "White", "gold": "Gold",
    "japanese gold": "Gold", "rose gold": "Gold", "silver": "Silver",
    "beige": "Beige", "brown": "Brown", "blue": "Blue", "red": "Red",
    "green": "Green", "gray": "Grey", "grey": "Grey", "pink": "Pink",
    "purple": "Purple", "orange": "Orange", "transparent": "Clear",
    "nude": "Beige", "multi": "Multicolour", "multicolor": "Multicolour",
    "tortoise": "Brown", "havana": "Brown", "leopard": "Brown",
    "burgundy": "Red", "navy": "Blue", "navy blue": "Blue",
    "yellow": "Yellow", "cream": "Beige",
}
FRAME_SHAPE_MAP = {
    "geometric": "Irregular", "rectangle": "Rectangular", "rectangular": "Rectangular",
    "oval": "Oval", "round": "Round", "cat eye": "Cat Eye", "cateye": "Cat Eye",
    "square": "Square", "aviator": "Aviator", "teardrop": "Pilot",
    "browline": "Browline", "butterfly": "Butterfly", "heart": "Heart",
    "oversized": "Oversized", "wayfarer": "Wayfarer", "wrap": "Wrap",
    "sport": "Sport", "shield": "Shield",
}
FRAME_MAT_MAP = {
    "metal": "Metal", "acetate": "Acetate", "tr90": "Plastic", "tr-90": "Plastic",
    "plastic": "Plastic", "nylon": "Plastic", "titanium": "Titanium",
    "aluminum": "Aluminium", "aluminium": "Aluminium",
    "stainless steel": "Stainless Steel",
}

# Trendyol-specific frame material map (TR90 → Plastic, mixed → Mixed)
FRAME_MAT_MAP_TY = {
    "metal": "Metal", "acetate": "Acetate", "tr90": "Plastic", "tr-90": "Plastic",
    "plastic": "Plastic", "nylon": "Plastic", "aluminum": "Aluminum",
    "titanium": "Titanium", "mixed": "Mixed", "wood": "Wooden", "bone": "Bone",
}
RIM_MAP = {
    "full rim": "Full Rim", "half rim": "Semi-Rimless",
    "rimless": "Rimless", "semi-rimless": "Semi-Rimless",
    "semirimless": "Semi-Rimless",
}
LENS_CAT_MAP = {
    "cat. 0": "Category 0", "cat. 1": "Category 1", "cat. 2": "Category 2",
    "cat. 3": "Category 3", "cat. 4": "Category 4",
    "category 0": "Category 0", "category 1": "Category 1",
    "category 2": "Category 2", "category 3": "Category 3",
}

GENDER_MAP_TY = {
    "women": "Female", "woman": "Female", "female": "Female",
    "men": "Male", "man": "Male", "male": "Male", "boys": "Male", "boy": "Male",
    "unisex": "Unisex", "kids": "Unisex", "girls": "Female", "girl": "Female",
}

FRAME_TYPE_MAP = {
    "full rim": "Framed", "half rim": "Framed",
    "rimless": "Frameless", "semi-rimless": "Framed", "semirimless": "Framed",
}

FRAME_COLOR_MAP_TY = {
    "japanese gold": "Gold", "gold": "Gold", "rose gold": "Rose gold", "silver": "Silver",
    "beige": "Beige", "brown": "Brown", "black": "Black", "blue": "Blue", "red": "Red",
    "green": "Green", "gray": "Gray", "grey": "Gray", "white": "Single Color",
    "pink": "Pink", "purple": "Purple", "orange": "Orange", "transparent": "Transparent",
    "nude": "Nude", "burgundy": "Burgundy", "navy blue": "Navy blue", "multi": "Multi-colored",
    "metallic": "Metallic", "bronze": "Bronze", "copper": "Copper",
}

WEB_COLOR_MAP = {
    "gold": "Gold-colored", "japanese gold": "Gold-colored", "rose gold": "Pink",
    "silver": "Silver-colored", "beige": "Beige", "brown": "Brown", "black": "Black",
    "blue": "Blue", "red": "Red", "green": "Green", "gray": "Gray", "grey": "Gray",
    "white": "White", "pink": "Pink", "purple": "Purple", "orange": "Orange",
    "transparent": "White", "nude": "Cream", "burgundy": "Bordeaux",
    "navy blue": "Dark blue", "metallic": "Metallic", "bronze": "Gold-colored",
    "copper": "Gold-colored", "multi": "Multicolor", "multicolor": "Multicolor",
}

LENS_FINISH_MAP = {
    "gradient": "Gradient", "mirrored": "Mirrored", "mat": "Mat", "matte": "Mat",
}

BRAND_DESC = {
    "30Sundays": (
        "30Sundays is a contemporary eyewear brand focused on stylish, everyday essentials that blend comfort, "
        "quality, and affordability. The brand is designed for modern lifestyles, offering versatile frames that "
        "suit both work and casual wear. With a strong focus on value and wearability, 30Sundays makes eye care "
        "fashionable and accessible."
    ),
    "Babamio": (
        "Babamio is a vibrant eyewear brand for children, dedicated to bringing bold color, playful designs, and "
        "UV protection to young adventurers. From beloved character franchises to everyday styles, each pair is "
        "crafted with durable, BPA-free materials that are safe and comfortable for kids. With Babamio, little ones "
        "can express their personality while keeping their eyes protected."
    ),
    "BlackOut": (
        "BlackOut is a bold contemporary eyewear brand designed for the modern woman. Combining precision-crafted "
        "frames with distinctive geometric silhouettes, BlackOut sunglasses deliver a confident statement in every "
        "look. Each pair offers full UV400 protection and is built from premium materials — making style and eye "
        "safety equally non-negotiable."
    ),
}

BULLETS = {
    "30Sundays": [
        "Timeless silhouettes inspired by heritage aesthetics, blending traditional elegance with modern minimal design suitable for Ramadan and festive wear.",
        "Lightweight, comfortable frames designed for all-day wear, ideal for long gatherings, evenings, and celebrations.",
        "Versatile color tones and refined finishes that effortlessly complement both ethnic and contemporary outfits.",
    ],
    "Babamio": [
        "Designed for young adventurers — durable, BPA-free frames built to handle active kids' everyday wear safely and comfortably.",
        "Full UV400 protection shields little eyes from harmful sun rays during outdoor play, sports, and everyday activities.",
        "Fun, vibrant colors and character-inspired designs that kids love, available in sizes crafted specifically for growing faces.",
    ],
    "BlackOut": [
        "Bold geometric silhouettes and premium materials deliver a confident, fashion-forward look for the modern woman.",
        "Full UV400 protection combined with lightweight metal and TR90 construction for comfort you can wear all day.",
        "Versatile colorways and distinctive frame finishes that transition effortlessly from casual outings to styled occasions.",
    ],
}

FIELDS = {
    "short_sku": ["Short SKU"],
    "product_name": ["Product Name"],
    "brand": ["Brand"],
    "gender": ["Gender"],
    "frame_material": ["Frame Material", "Front Material"],
    "frame_shape": ["Frame Shape"],
    "rim_type": ["Rim Type"],
    "lens_color": ["Lens Color"],
    "lens_finish": ["Lens Finish"],
    "lens_uv": ["Lens UV Rating"],
    "lens_category": ["Lens Category Rating"],
    "frame_color_eyewa": ["Frame Color (Eyewa)", "Front Color (Specific)"],
    "frame_color_group": ["Frame Color (Group)", "Front Color Group"],
    "lens_size": ["Lens Size"],
    "bridge": ["Bridge"],
    "temple_length": ["Temple Length"],
    "age_group": ["Age Group"],
    "srp_aed": ["Eyewa SRP (AED)"],
    "srp_sar": ["Eyewa SRP (SAR)"],
}

SHEET_CONFIG = [
    ("30Sundays SG", 0, "Eyewa SKU / ERP"),
    ("Babamio", 1, "Eyewa SKU / ERP Search Name"),
    ("Old Babamio", 1, "Eyewa SKU / ERP Search Name"),
    ("BlackOut", 1, "Eyewa SKU / ERP Search Name"),
]

# ─── Helper functions ─────────────────────────────────────────────────────────
def mv(mapping, val):
    return mapping.get(str(val).strip().lower()) if val else None

def map_frame_mat(val):
    if not val:
        return None
    v = str(val).strip()
    if "+" in v or "/" in v:
        for p in [x.strip().lower() for x in v.replace("/", "+").split("+")]:
            mm = FRAME_MAT_MAP.get(p)
            if mm:
                return mm
        return "Metal"
    return FRAME_MAT_MAP.get(v.lower())

def map_lens_type(uv_val, finish_val):
    if finish_val and "polariz" in str(finish_val).lower():
        return "Polarized"
    if finish_val and "mirror" in str(finish_val).lower():
        return "Mirrored"
    return "UV Protection"

def model_name_from_title(title, brand):
    t = str(title).strip() if title else ""
    return t[len(brand) + 1:] if t.startswith(brand + " ") else t

def safe_num(v):
    if v is None:
        return None
    try:
        return int(float(str(v)))
    except Exception:
        return None

def build_col_map(headers, fields):
    m = {}
    hl = {str(h).strip().lower(): i for i, h in enumerate(headers) if h}
    for fld, aliases in fields.items():
        for a in aliases:
            if a.lower() in hl:
                m[fld] = hl[a.lower()]
                break
    return m

# ─── Data loading (cached) ────────────────────────────────────────────────────
@st.cache_data(show_spinner="Loading RAP sheet…")
def load_rap(rap_bytes):
    wb = openpyxl.load_workbook(io.BytesIO(rap_bytes), data_only=True)
    rap_index = {}
    model_index = {}
    for sheet_name, hrow, sku_col_name in SHEET_CONFIG:
        if sheet_name not in wb.sheetnames:
            continue
        ws = wb[sheet_name]
        rows = list(ws.iter_rows(values_only=True))
        headers = list(rows[hrow])
        sku_col = next(
            (i for i, h in enumerate(headers) if h and str(h).strip() == sku_col_name), None
        )
        if sku_col is None:
            sku_col = next(
                (i for i, h in enumerate(headers) if h and "Eyewa SKU" in str(h)), None
            )
        col_map = build_col_map(headers, FIELDS)
        for row in rows[hrow + 1:]:
            if sku_col is None or sku_col >= len(row):
                continue
            sku = row[sku_col]
            if not sku or not isinstance(sku, str) or sku[0] not in ("s", "S"):
                continue
            if sku in rap_index:
                continue
            rec = {"_sheet": sheet_name, "_sku": sku}
            for fld, cidx in col_map.items():
                rec[fld] = row[cidx] if cidx < len(row) else None
            rap_index[sku] = rec
            parts = sku.split("-")
            if len(parts) >= 3:
                model_index.setdefault((parts[1], parts[2]), []).append(sku)
    return rap_index, model_index

@st.cache_data(show_spinner="Loading images…")
def load_images(img_bytes):
    img_index = {}
    content = img_bytes.decode("utf-8", errors="replace")
    reader = csv.reader(content.splitlines())
    hdrs = next(reader)
    img_cols = [i for i, h in enumerate(hdrs) if "Image" in h]
    for row in reader:
        sku = row[1] if len(row) > 1 else ""
        if not sku:
            continue
        imgs = list(OrderedDict.fromkeys(
            r for i in img_cols if i < len(row) for r in [row[i]] if r
        ))
        img_index[sku] = imgs
    return img_index

def get_rap(sku, rap_index, model_index):
    if sku in rap_index:
        return rap_index[sku]
    parts = sku.split("-")
    if len(parts) >= 3:
        cands = model_index.get((parts[1], parts[2]), [])
        if cands:
            return rap_index[cands[0]]
    return None

def get_images(sku, img_index, rap_index=None, model_index=None):
    if sku in img_index:
        return img_index[sku]
    parts = sku.split("-")
    if len(parts) >= 3:
        model, brand = parts[2], parts[1]
        for k, v in img_index.items():
            if model in k and brand in k:
                return v
    return []

# ─── Row builders ─────────────────────────────────────────────────────────────
def map_frame_mat_ty(val):
    """Trendyol frame material mapping: TR90→Plastic, combined→Mixed."""
    if not val:
        return None
    v = str(val).strip()
    if "+" in v or "/" in v:
        return "Mixed"
    return FRAME_MAT_MAP_TY.get(v.lower(), v[:20] if v else None)

def map_lens_uv(val):
    if not val:
        return "UV400"
    v = str(val).strip().lower()
    return "Polarized" if "polariz" in v else "UV400"

def build_trendyol_row(sku, rap_index, model_index, img_index):
    d = get_rap(sku, rap_index, model_index)
    imgs = get_images(sku, img_index)
    if not imgs and d:
        imgs = get_images(d.get("_sku", sku), img_index)

    parts = sku.split("-")
    brand_code = parts[1] if len(parts) > 1 else ""
    brand_from_sku = {"30s": "30Sundays", "bab": "Babamio", "bla": "BlackOut"}.get(brand_code, "")
    img_list = (imgs + [None] * 8)[:8]

    if d is None:
        return (
            sku, None, brand_from_sku, "379", brand_from_sku,
            BRAND_DESC.get(brand_from_sku, ""), 0, 0, 0, None, "0",
            *img_list,
            None, None, None, None, None, None, None, None, None,
            None, None, None, None, None, None, None,
        )

    brand      = str(d.get("brand") or brand_from_sku)
    title      = str(d.get("product_name") or brand)
    short_sku  = str(d.get("short_sku") or "") or None
    srp_aed    = safe_num(d.get("srp_aed")) or 0
    gender     = mv(GENDER_MAP_TY, str(d.get("gender") or "").lower()) or "Female"

    raw_fc     = str(d.get("frame_color_eyewa") or "")
    raw_fcg    = str(d.get("frame_color_group") or "")
    frame_color = FRAME_COLOR_MAP_TY.get(raw_fc.lower()) or (raw_fc[:20] if raw_fc else None)
    web_color   = WEB_COLOR_MAP.get(raw_fcg.lower()) or (raw_fcg[:20] if raw_fcg else None)

    lens_color  = str(d.get("lens_color") or "") or None
    # Frame shape: use raw RAP value (no mapping — Trendyol accepts the raw string)
    frame_shape = str(d.get("frame_shape") or "") or None
    frame_mat   = map_frame_mat_ty(d.get("frame_material"))
    frame_type  = FRAME_TYPE_MAP.get(str(d.get("rim_type") or "").lower())
    glass_type  = LENS_FINISH_MAP.get(str(d.get("lens_finish") or "").lower())
    lens_mat    = map_lens_uv(d.get("lens_uv"))
    ls_raw      = d.get("lens_size")
    lens_width  = str(int(float(ls_raw))) if ls_raw and str(ls_raw).replace(".", "").isdigit() else None
    desc        = BRAND_DESC.get(brand, "")

    # 37 columns — exact match to Trendyol template column order:
    # [00] Barcode, [01] ModelCode, [02] Brand, [03] CategoryId, [04] Title,
    # [05] Description, [06] OriginalPrice, [07] SalePrice, [08] Stock,
    # [09] StockCode, [10] VatRate, [11-18] Images x8, [19] HandlingTime,
    # [20] Lens Material, [21] Pattern, [22] Frame Type, [23] Glass Type,
    # [24] Additional Feature, [25] Lens Color, [26] Frame Color, [27] Gender,
    # [28] Age Group, [29] Color, [30] Web Color, [31] Origin,
    # [32] Lens Width, [33] Frame Shape, [34] Frame Material,
    # [35] Instructions, [36] Material Composition
    return (
        sku, short_sku, brand, "379", title,
        desc, srp_aed, srp_aed, 0, None, "0",
        *img_list,
        None, lens_mat, None, frame_type, glass_type,
        None, lens_color, frame_color, gender,
        None, frame_color, web_color, None,
        lens_width, frame_shape, frame_mat,
        None, None,
    )

def build_noon_row(sku, rap_index, model_index, img_index):
    d = get_rap(sku, rap_index, model_index)
    imgs = get_images(sku, img_index)
    if not imgs and d:
        imgs = get_images(d.get("_sku", sku), img_index)
    img_list = (imgs + [None] * 7)[:7]

    parts = sku.split("-")
    brand_code = parts[1] if len(parts) > 1 else ""
    brand_from_sku = {"30s": "30Sundays", "bab": "Babamio", "bla": "BlackOut"}.get(brand_code, "")

    if d is None:
        row = [
            "Eyewear", "Eyewear", "Sunglasses", brand_from_sku, brand_from_sku, sku, "", "",
            "New", "New", None, None, None, None, None, None, None, None, None, None,
            None, None, None, None, None, None, None, None, None, None, None,
        ]
        row += list(BULLETS.get(brand_from_sku, ["", "", ""])[:3]) + [None] * 9
        row += img_list
        row += [None, None, "China", 15, "Centimeter", 6, "Centimeter", 12, "Centimeter",
                0.2, "Kilogram", 13, "Centimeter", 4, "Centimeter", 10, "Centimeter",
                None, None, None, None, None, None]
        return row

    brand = str(d.get("brand") or brand_from_sku)
    title = str(d.get("product_name") or brand)
    short_sku = str(d.get("short_sku") or "")
    mn = model_name_from_title(title, brand)
    srp_aed = safe_num(d.get("srp_aed"))
    srp_sar = safe_num(d.get("srp_sar"))
    dept = mv(DEPT_MAP, str(d.get("gender") or "").lower()) or "Unisex"
    raw_lc = str(d.get("lens_color") or "")
    raw_fc = str(d.get("frame_color_eyewa") or "")
    lens_col = raw_lc or None
    lens_col_family = FRAME_COLOR_FAMILY.get(raw_lc.lower()) or raw_lc or None
    frame_col = raw_fc or None
    frame_col_family = (
        FRAME_COLOR_FAMILY.get(raw_fc.lower())
        or FRAME_COLOR_FAMILY.get(str(d.get("frame_color_group") or "").lower())
        or None
    )
    shape = mv(FRAME_SHAPE_MAP, str(d.get("frame_shape") or "").lower())
    rim = mv(RIM_MAP, str(d.get("rim_type") or "").lower())
    fmat = map_frame_mat(d.get("frame_material"))
    lens_cat = mv(LENS_CAT_MAP, str(d.get("lens_category") or "").lower()) or "Category 3"
    lens_t = map_lens_type(d.get("lens_uv"), d.get("lens_finish"))
    bridge = safe_num(d.get("bridge"))
    temple = safe_num(d.get("temple_length"))
    ls = safe_num(d.get("lens_size"))
    bullets = BULLETS.get(brand, ["", "", ""])

    row = [
        "Eyewear", "Eyewear", "Sunglasses", brand, title, sku, mn, short_sku,
        "New", "New", dept,
        lens_col, lens_col_family, frame_col, frame_col_family, shape,
        bridge, "Millimeter" if bridge else None,
        temple, "Millimeter" if temple else None,
        ls, "Millimeter" if ls else None,
        None, None, lens_t, None, fmat, None, rim, lens_cat, None,
    ]
    row += [
        bullets[0] if len(bullets) > 0 else None,
        bullets[1] if len(bullets) > 1 else None,
        bullets[2] if len(bullets) > 2 else None,
    ]
    row += [None] * 9
    row += img_list
    row += [None, None, "China"]
    row += [15, "Centimeter", 6, "Centimeter", 12, "Centimeter", 0.2, "Kilogram"]
    row += [13, "Centimeter", 4, "Centimeter", 10, "Centimeter"]
    row += [srp_aed, srp_sar, None, None, None, None]
    return row

# ─── Template generators ───────────────────────────────────────────────────────
def generate_trendyol(skus, rap_index, model_index, img_index, template_bytes):
    wb = openpyxl.load_workbook(io.BytesIO(template_bytes))
    # Find the data sheet — try exact name first, then keyword match, then first sheet
    target = "Enter your product information"
    if target not in wb.sheetnames:
        target = next(
            (n for n in wb.sheetnames if "product" in n.lower() or "data" in n.lower()),
            wb.sheetnames[0],
        )
    ws = wb[target]
    # Find the real last row with data (ignore ghost/formatted-only rows)
    last_data_row = 0
    for i, row in enumerate(ws.iter_rows(values_only=True), start=1):
        if any(v is not None for v in row):
            last_data_row = i
    start_row = last_data_row + 1
    for i, sku in enumerate(skus):
        row_data = list(build_trendyol_row(sku, rap_index, model_index, img_index))
        for col, val in enumerate(row_data, start=1):
            ws.cell(row=start_row + i, column=col, value=val)
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue(), target, len(skus)

def generate_noon(skus, rap_index, model_index, img_index, template_bytes):
    wb = openpyxl.load_workbook(io.BytesIO(template_bytes))
    ws = wb["template_data"]
    for sku in skus:
        ws.append(build_noon_row(sku, rap_index, model_index, img_index))
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()

# ─── UI ────────────────────────────────────────────────────────────────────────
with st.sidebar:
    st.header("📁 Upload Data Files")
    rap_file = st.file_uploader("RAP Sheet (.xlsx)", type=["xlsx"], key="rap")
    img_file = st.file_uploader("Eyewa Images (.csv)", type=["csv"], key="img")
    st.divider()
    st.header("📋 Templates")
    ty_template = st.file_uploader("Trendyol Template (.xlsx)", type=["xlsx"], key="ty_tmpl")
    noon_template = st.file_uploader("Noon Template (.xlsx)", type=["xlsx"], key="noon_tmpl")
    st.divider()
    st.caption("Upload files once — they're cached for the session.")

col1, col2 = st.columns([2, 1])

with col1:
    st.subheader("SKU List")
    sku_text = st.text_area(
        "Paste SKU codes here (one per line)",
        height=300,
        placeholder="sw100104-30s-000490-0104-51\nsw30107-30s-000491-0107p-53\n...",
    )

with col2:
    st.subheader("Generate")
    platform = st.radio(
        "Platform",
        ["Trendyol", "Noon", "Both"],
        horizontal=False,
    )

    generate_btn = st.button("⚡ Generate Template", type="primary", use_container_width=True)

# ─── Processing ────────────────────────────────────────────────────────────────
if generate_btn:
    # Parse SKUs
    raw_lines = [l.strip() for l in sku_text.strip().splitlines() if l.strip()]
    skus = [l.split()[0] for l in raw_lines if l]
    skus = [s for s in skus if s.startswith(("s", "S"))]

    if not skus:
        st.error("No valid SKUs found. Make sure each line starts with a SKU (e.g. sw...).")
        st.stop()

    if not rap_file:
        st.error("Please upload the RAP Sheet in the sidebar.")
        st.stop()

    if not img_file:
        st.error("Please upload the Eyewa Images CSV in the sidebar.")
        st.stop()

    # Load data (seek to start before reading — Streamlit file pointer may be at end)
    rap_file.seek(0)
    img_file.seek(0)
    rap_index, model_index = load_rap(rap_file.read())
    img_index = load_images(img_file.read())

    # Stats
    no_rap = [s for s in skus if get_rap(s, rap_index, model_index) is None]
    no_img = [
        s for s in skus
        if not get_images(s, img_index)
        and not get_images(
            (get_rap(s, rap_index, model_index) or {}).get("_sku", s), img_index
        )
    ]

    st.success(f"✅ {len(skus)} SKUs parsed — {len(no_rap)} missing from RAP, {len(no_img)} missing images")

    if no_rap:
        with st.expander(f"⚠️ {len(no_rap)} SKUs not found in RAP sheet"):
            st.code("\n".join(no_rap))

    if no_img:
        with st.expander(f"⚠️ {len(no_img)} SKUs missing images"):
            st.code("\n".join(no_img))

    # Generate files
    files_to_zip = {}

    if platform in ("Trendyol", "Both"):
        if not ty_template:
            st.warning("Trendyol template not uploaded — skipping Trendyol.")
        else:
            with st.spinner("Building Trendyol template…"):
                try:
                    ty_template.seek(0)
                    ty_bytes, ty_sheet, ty_rows = generate_trendyol(
                        skus, rap_index, model_index, img_index, ty_template.read()
                    )
                    files_to_zip["Trendyol_Filled.xlsx"] = ty_bytes
                except Exception as e:
                    st.error(f"Trendyol generation failed: {e}")

    if platform in ("Noon", "Both"):
        if not noon_template:
            st.warning("Noon template not uploaded — skipping Noon.")
        else:
            with st.spinner("Building Noon template…"):
                try:
                    noon_template.seek(0)
                    noon_bytes = generate_noon(skus, rap_index, model_index, img_index, noon_template.read())
                    files_to_zip["Noon_Filled.xlsx"] = noon_bytes
                except Exception as e:
                    st.error(f"Noon generation failed: {e}")

    if files_to_zip:
        if len(files_to_zip) == 1:
            # Single file — download directly
            fname, fdata = next(iter(files_to_zip.items()))
            st.download_button(
                f"⬇️ Download {fname}",
                data=fdata,
                file_name=fname,
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True,
                type="primary",
            )
        else:
            # Multiple files — zip them
            zip_buf = io.BytesIO()
            with zipfile.ZipFile(zip_buf, "w", zipfile.ZIP_DEFLATED) as zf:
                for fname, fdata in files_to_zip.items():
                    zf.writestr(fname, fdata)
            st.download_button(
                "⬇️ Download All Templates (ZIP)",
                data=zip_buf.getvalue(),
                file_name="Eyewa_Templates.zip",
                mime="application/zip",
                use_container_width=True,
                type="primary",
            )
